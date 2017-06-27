import os
import xlwings as xw
import datetime as dt
import openpyxl as oxl
from openpyxl import Workbook
from collections import OrderedDict
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment
from openpyxl.writer.write_only import WriteOnlyCell

'''
Program Constants
'''

WORKBOOK_PATH = 'test-book.xlsm'

HOME_SHEET = 'Home'
RTE_SHEET = 'RTE Spreadsheet'
SM_SHEET = 'SM Spreadsheet'
MM_SHEET = 'MM Spreadsheet'

OUTPUT_TAB = 'RTE-SM Compare'

RTE_VALIDATION = 'Production MM'
SM_VALIDATION = 'SM Version'
MM_VALIDATION = 'Production MM'

OPERATION_START = 'Full Oper Num'

SAVE_ROOT_PATH = ''
HEADER_BUFFER = 3
NUM_OUTPUT_COLS = 30
COL_A = 0
COL_B = 1

RED = 'FFFF0000'
RED_RGB = (255,0,0)
RTE_HEX = 'FFF0F0F0'
SM_HEX = 'FFFFFF99'
ORANGE_DIFF_HEX = 'FFFF9966'
BLUE_DIFF_HEX = 'FF00B0F0'
HOME_BLUE_RGB = (146, 205, 220)
YELLOW_HEX = 'FFFF00'
ORANGE_WARNING_RGB = (255, 192, 0)

RTE_FILL = PatternFill(patternType='solid', fgColor=RTE_HEX)
SM_FILL = PatternFill(patternType='solid', fgColor=SM_HEX)
ORANGE_DIFF_FILL = PatternFill(patternType='solid', fgColor=ORANGE_DIFF_HEX)
BLUE_DIFF_FILL = PatternFill(patternType='solid', fgColor=BLUE_DIFF_HEX)

HEADER_FONT = Font(name='Calibri', bold=True, size=14)

ALL_BORDER = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

CENTER_ALIGN = Alignment(horizontal='center')

BLUE = 'Blue'
ORANGE = 'Orange'



#pass in cell, return .value as string with no leading/trailing whitespace
def stringify(cell):
    return str(cell.value).strip()

'''
returns True if the openpyxl cell has a value, False otherwise
'''
def has_value(cell):
    return cell.value != '' and cell.value != None

'''
given a report_id, extract the route ID
'''
def get_route_id(report_id):
    try:
        '''
        Expected report_ids:
        Flow Report (SM Version): [ROUTE_ID], [PRODUCT_ID]
        Flow Report (Production MM): [ROUTE_ID], [PRODUCT_ID]
        '''
        return report_id.split(':')[1].strip().split(',')[0]
    except:
        return 'ErrorReadingRouteID'

'''
given a report_id, extract the product ID
'''
def get_product_id(report_id):
    try:
        '''
        Expected report_ids:
        Flow Report (SM Version): [ROUTE_ID], [PRODUCT_ID]
        Flow Report (Production MM): [ROUTE_ID], [PRODUCT_ID]
        '''
        return report_id.split(':')[1].strip().split(',')[1].strip()
    except:
        return 'ErrorReadingProduct'

'''
saves output_worksheet to the generated path & filename
'''
def save_output():
    file_name = generate_filename(rte_route)
    output_workbook.save(file_name)
    format_output(file_name)
    xw.Book(file_name)

'''
initializes the WriteOnlyWorkbook that is used for output
creates approproate sheets for output and user post processing verification
'''
def init_write_only_output():
    workbook = Workbook(write_only=True)
    workbook.create_sheet(OUTPUT_TAB)
    for option in get_output_options():
        workbook.create_sheet(option)

    for sheet in workbook.sheetnames:
        workbook[sheet].sheet_properties.tabColor = YELLOW_HEX
    
    for i in range(HEADER_BUFFER):
        workbook[OUTPUT_TAB].append([''])
    
    return workbook

'''
given a route, create a filename. 
    [SAVE_ROOT_PATH]/
        [ROUTE_ID]_RTE-SM-Compare_[YEAR-MONTH-DATE] [HOUR-MINUTE-SECOND]_[WINDOWS_ID]
creates user directory if not already existing
@param route: route to create filename for
'''
def generate_filename(route):
    uid = os.getlogin()
    save_path = SAVE_ROOT_PATH + uid
    timestamp = str(dt.datetime.now())[:-7].replace(":", "-")
    file_name = save_path + '/' + homesheet.range('E17').value + '_' + route.get_route_id() + '_RTE-SM-Compare_' + timestamp + '_' + uid + '.xlsx'
    
    if not os.path.exists(save_path):
        os.mkdir(save_path)
        
    return file_name

'''
creates a WriteOnlyCell and returns it to the caller
@param value: value for cell
@param route_type: fills a color based on route_type [RTE, SM]
@param diff: fills a color based on diff type [ORANGE, BLUE]
'''
def create_cell(value, route_type, diff=None):
    cell = WriteOnlyCell(ws=output_worksheet, value=value)
    if route_type == 'RTE':
        if diff == ORANGE and has_value(cell):
            cell.fill = ORANGE_DIFF_FILL
        elif diff == BLUE and has_value(cell):
            cell.fill = BLUE_DIFF_FILL
        else:
            cell.fill = RTE_FILL
    elif route_type == 'SM':
        if diff == ORANGE and has_value(cell):
            cell.fill = ORANGE_DIFF_FILL
        elif diff == BLUE and has_value(cell):
            cell.fill = BLUE_DIFF_FILL
        else:
            cell.fill = SM_FILL
    return cell

def part_of_change(cell):
    try:
        if cell.fill.fgColor.rgb != '00000000':
            return True
    except AttributeError:
        return False
    return False

'''
given a route_type and operation, render the extra operation to the output_worksheet
@param route_type: responsible for determining type of output
@param operation: operation to output
'''
def render_extra_operation(route_type, operation):
    if type(operation) != Operation:
        print('[X] render_extra_operation recieved bad operation.')
        return False
    
    write_header()
    
    if route_type == 'RTE':
        for row in operation.get_operation_as_output('RTE', extra=True, difftype=ORANGE):
            output_worksheet.append(row)
        print('\t[+]Output extra RTE operation:', operation)
    elif route_type == 'SM':
        for row in operation.get_operation_as_output('SM', extra=True, difftype=BLUE):
            output_worksheet.append(row)
        print('\t[+]Output extra SM operation', operation)
    else:
        print('[X]render_extra_operation recieved bad route_type')

'''
compares the given routes. RTE and SM routes must be present at a minimum
    if MM route is present, logic will account for the sheet automatically
@return: openpyxl WriteOnlyWorksheet to be saved output containing the route differences  
'''

def compare_routes(rte_route, sm_route, mm_route=None):
    
    #Operation numbers stored as list
    rte_operation_nums = rte_route.get_operation_nums()
    sm_operation_nums = sm_route.get_operation_nums()
    
    
    
    #Parallel cursors to account for each route potentially having extra (unique) operations
    rte_cursor = 0
    sm_cursor = 0
    
    
    
    #Loop will run until the cursor for each route is at the route's last operation
    comparing = True
    
    while comparing:
        #print('RTE cursor:', rte_cursor, 'SM cursor:', sm_cursor)
        '''
        Handling extra operations in either the RTE or SM sheet
        If an SM sheet is present, check against sheet to see if can be neutralized
        '''
        
        rte_op_num_str = rte_operation_nums[rte_cursor]
        sm_op_num_str = sm_operation_nums[sm_cursor]
        
        
        #If the current operation numbers do not equal each other, either the RTE or SM has an extra operation
        if rte_operation_nums[rte_cursor] != sm_operation_nums[sm_cursor]:
            print('[X]Operation misalign. RTE: ', rte_operation_nums[rte_cursor], 'SM:', sm_operation_nums[sm_cursor])
            
            #Store our operation numbers as floats for number comparisons
            
            
            rte_op_num = float(rte_operation_nums[rte_cursor])
            sm_op_num = float(sm_operation_nums[sm_cursor])
            
            #If the RTE contains an operation that is not in the SM 
            #AND the SM contains and operation that is not in the RTE
            #Consider a complete mismatch senario where both operations are considered extra
            #Ignore cursors
            if rte_route.has_operation(sm_op_num_str) == False and sm_route.has_operation(rte_op_num_str) == False:
                #print('Complete mismatch!')
                print('\t[X]Extra RTE operation', rte_op_num_str)
                render_difference(rte_route.operations[rte_op_num_str], None)
                print('\t[X]Extra SM operation', sm_op_num_str)
                render_difference(None, sm_route.operations[sm_op_num_str])
            
            
            
            #If the SM operation number is ahead of the RTE, the RTE has an extra operation
            elif rte_op_num < sm_op_num:
                #Increment the RTE cursor until the operation numbers are realigned
                while rte_op_num < sm_op_num:
                    print('\t[X]Extra RTE operation', rte_op_num_str)
                    render_difference(rte_route.operations[rte_op_num_str], None)
                    #render_difference(rte_route.operations[rte_op_num_str], None)
                    rte_cursor += 1                    
                    rte_op_num = float(rte_operation_nums[rte_cursor])
                    
                print('[+]Operations realigned. RTE: ', rte_operation_nums[rte_cursor], 'SM:', sm_operation_nums[sm_cursor])
                #Now need to compare realigned operaions
                
                rte_op_num_str = rte_operation_nums[rte_cursor]
                sm_op_num_str = sm_operation_nums[sm_cursor]
                print('[+]Comparing RTE:', rte_route.operations[rte_op_num_str], 'SM:', sm_route.operations[sm_op_num_str])
                if rte_route.operations[rte_op_num_str] != sm_route.operations[sm_op_num_str]:
                #Operation diff
                    if (rte_route.operations[rte_op_num_str].part_of_change == False and 
                        mm_route != None and
                        mm_route.has_operation(rte_op_num_str) and 
                        sm_route.operations[rte_op_num_str] == mm_route.operations[rte_op_num_str]):
                        print('\t[+]Operation difference neutralized by MM sheet.:')
                    else:
                        render_difference(rte_route.operations[rte_op_num_str], sm_route.operations[sm_op_num_str])
                else:
                    print('\t[+]Operations equal.:', rte_route.operations[rte_op_num_str], 'SM:', sm_route.operations[sm_op_num_str])
                
                
            #If the RTE operation number is ahead of the SM spreadsheet, the SM sheet has an extra operation
            elif rte_op_num > sm_op_num:
                #Increment the SM cursor until the operation numbers are realigned
                while rte_op_num > sm_op_num:
                    print('\tExtra SM operation', sm_op_num_str)
                    render_difference(None, sm_route.operations[sm_op_num_str])
                    #render_difference(None, sm_route.operations[sm_op_num_str])
                    sm_cursor += 1
                    sm_op_num = float(sm_operation_nums[sm_cursor])
                    
                print('[+]Operations realigned. RTE: ', rte_operation_nums[rte_cursor], 'SM:', sm_operation_nums[sm_cursor])
                #Now need to compare realigned operaions
                
                rte_op_num_str = rte_operation_nums[rte_cursor]
                sm_op_num_str = sm_operation_nums[sm_cursor]
                if rte_route.operations[rte_op_num_str] != sm_route.operations[sm_op_num_str]:
                #Operation diff
                    if (rte_route.operations[rte_op_num_str].part_of_change == False and 
                        mm_route != None and
                        mm_route.has_operation(rte_op_num_str) and 
                        sm_route.operations[rte_op_num_str] == mm_route.operations[rte_op_num_str]):
                        print('\t[+]Operation difference neutralized by MM sheet.:')
                    else:
                        render_difference(rte_route.operations[rte_op_num_str], sm_route.operations[sm_op_num_str])
                else:
                    print('\t[+]Operations equal.:', rte_route.operations[rte_op_num_str], 'SM:', sm_route.operations[sm_op_num_str])
    
            #Safe check - if control falls here means that somehow the operation numbers are not less than/ greater than eachother AND not equal to eachother...
            else:
                print('[X]You should not be here, check operation alignment loop.')
        #Operation numbers are equal, proceed to normal operation comparison
        else:
            print('[+]Comparing RTE:', rte_route.operations[rte_op_num_str], 'SM:', sm_route.operations[sm_op_num_str])
            #print('[+]Operation numbers equal. RTE:', rte_operation_nums[rte_cursor], 'SM:', sm_operation_nums[sm_cursor])
            if rte_route.operations[rte_op_num_str] != sm_route.operations[sm_op_num_str]:
                #Operation diff
                if (rte_route.operations[rte_op_num_str].part_of_change == False and 
                        mm_route != None and
                        mm_route.has_operation(rte_op_num_str) and 
                        sm_route.operations[rte_op_num_str] == mm_route.operations[rte_op_num_str]):
                        print('\t[+]Operation difference neutralized by MM sheet.:')
                else:
                    render_difference(rte_route.operations[rte_op_num_str], sm_route.operations[sm_op_num_str])
            else:
                print('\t[+]Operations equal.:', rte_route.operations[rte_op_num_str], 'SM:', sm_route.operations[sm_op_num_str])
        #Increment our loop control variables
        rte_cursor += 1
        sm_cursor += 1
        
        
        #If the RTE or SM cursors are at the end of the routes, we are done comparing
        if rte_cursor == rte_route.get_num_operations() or sm_cursor == sm_route.get_num_operations():
            comparing = False
    
    #If either route is not equal to its cursor, it has extra trailing operations. 
    #For example, RTE-SM match for operations 1-10, but SM has operations 11-15, where RTE only has 1-10
    #Can assume these operations are equal
    while rte_cursor != rte_route.get_num_operations():
        rte_op_num_str = rte_operation_nums[rte_cursor]
        print('[X]Extra RTE operation', rte_op_num_str)
        render_difference(rte_route.operations[rte_op_num_str], None)
        rte_cursor += 1   
    while sm_cursor != sm_route.get_num_operations():
        sm_op_num_str = sm_operation_nums[sm_cursor]
        print('[X]Extra SM operation', sm_op_num_str)
        render_difference(None, sm_route.operations[sm_op_num_str])
        sm_cursor += 1 

def render_difference(rte_operation, sm_operation):
    print('\t[+]Writing operation difference.')
    
    write_header()
    row_counter = 0
    if type(rte_operation) == Operation and type(sm_operation) == Operation:
        for rte_row, sm_row in zip(rte_operation.rows, sm_operation.rows):
            temp_row = []
            for rte_cell, sm_cell in zip(rte_row, sm_row):
                if rte_cell.value != sm_cell.value and part_of_change(rte_cell):
                    temp_row.append(create_cell(rte_cell.value, 'RTE', ORANGE))
                    temp_row.append(create_cell(sm_cell.value, 'SM', ORANGE))
                elif rte_cell.value != sm_cell.value and not part_of_change(rte_cell):
                    temp_row.append(create_cell(rte_cell.value, 'RTE', BLUE))
                    temp_row.append(create_cell(sm_cell.value, 'SM', BLUE))
                else:
                    temp_row.append(create_cell(rte_cell.value, 'RTE'))
                    temp_row.append(create_cell(sm_cell.value, 'SM'))
            row_counter += 1
            output_worksheet.append(temp_row)
    
    #print('Row Counter:', row_counter, 'RTE Rows', len(rte_operation.rows), 'SM Rows', len(sm_operation.rows))
    #Extra RTE line
    while rte_operation != None and len(rte_operation.rows) > row_counter:
        temp_row = []
        row = [cell for cell in rte_operation.rows[row_counter]]
        for cell in row:
            if part_of_change(cell):
                temp_row.append(create_cell(cell.value, 'RTE', ORANGE))
                temp_row.append(create_cell('', 'SM'))
            else:
                temp_row.append(create_cell(cell.value, 'RTE', BLUE))
                temp_row.append(create_cell('', 'SM'))
        output_worksheet.append(temp_row)
        row_counter += 1
                
    while sm_operation != None and len(sm_operation.rows) > row_counter:
        temp_row = []
        row = [cell for cell in sm_operation.rows[row_counter]]
        for cell in row:
            if part_of_change(cell):
                temp_row.append(create_cell(cell.value, 'RTE', ORANGE))
                temp_row.append(create_cell('', 'SM'))
            else:
                temp_row.append(create_cell(cell.value, 'RTE', BLUE))
                temp_row.append(create_cell('', 'SM'))
        output_worksheet.append(temp_row)        
        row_counter += 1

'''
@OrderedDict: Used as main data structure to store operations in the form
    (Key, Value)    ->    (Operation Number, Operation Class) 
'''

'''
@Route: Stores all operations for a given route.
'''

def get_output_options():

    labels = xw.Range('H2:H9').value
    flags = xw.Range('I2:I9').value
    return [label for label, flag in zip(labels, flags) if flag]


class Route():
    
    '''
    @operations: Stores all operations for given route. Incoming values will be Operation type.
    @route_id: ID of the route pulled from the report_id
    @route_type: Will be either RTE (submitted), SM (staged chages), MM (production values)
    '''
    def __init__(self):
        self.operations = OrderedDict()
        self.route_id = None
        self.product_id = None
        self.route_type = None
        
    '''
    @return: returns the route id if not None, else 'NoSetRouteID' as error message
    '''
    def __str__(self):
        if self.route_id != None and self.route_type != None:
            return 'Route ID: ' + self.route_id + ', ' + self.product_id + ' | '  + self.route_type + ' | No. of Operations: ' + str(self.get_num_operations())  
        else:
            return 'NoSetRouteID'
    
    '''
    adds an operation to the route
    @param operation: will be of Operation type; operation to be added to route
        assuming operations are in sequential order
    '''
    def add_operation(self, operation):
        self.operations[str(operation)] = operation
        
    '''
    returns the route's operations to the caller
    @return: returns the operations of the route
    '''
    def get_operations(self):
        return self.operations
    
    '''
    sets the route's route ID
    @param route_id: ID of the current route
    '''
    def set_route_id(self, route_id):
        self.route_id = route_id
    
    '''
    sets the route's product ID
    @param product_id: product ID of the route
    '''
    def set_product_id(self, product_id):
        self.product_id = product_id
        
    '''
    sets the route's route type
    @param route_type: type of the current route
    '''
    def set_route_type(self, route_type):
        self.route_type = route_type
    
    '''
    returns total number of operations to the caller
    @return: int of total number of operations
    '''
    def get_num_operations(self):
        return len(self.operations)
    
    '''
    returns a list of the route's operation numbers (OrderedDict's keys)
    @return: returns a full list of operation numbers
    '''
    def get_operation_nums(self):
        return list(self.operations.keys())
    
    '''
    returns the last operation of the route to the caller
    @return: returns the last operation to the caller
    '''
    def get_last_operation(self):
        return self.operations[self.get_operation_nums()[self.get_num_operations()-1]]
    
    '''
    returns true if the route has the given operation, false otherwise
    @param operation: operation to look for
    '''
    def has_operation(self, operation):
        try:
            self.operations[operation]
            return True
        except KeyError:
            return False
        
    '''
    returns the route's route ID to the caller
    '''
    def get_route_id(self):
        return self.route_id
    
    '''
    returns the route's product ID to the caller
    '''
    def get_product_id(self):
        return self.product_id
    
class Operation():
    
    '''
    constructor
    '''
    def __init__(self):
        self.rows = []
        self.flagged_for_removal = False
        self.part_of_change = False
        
    '''
    returns the current operation's full operation number 
        'XXXX.XXXX'
    @return: operation's operation number
    '''
    def __str__(self):
        return self.fix_operation_no(self.rows[0][1].value)
    
    def __eq__(self, other_op):
        
        if type(other_op) != Operation:
            return False
        
        #Get operation as 2D array
        operation_as_list = self.get_operation_as_list()
        other_operation_as_list = other_op.get_operation_as_list()
        
        #Assume not the same if lengths are different 
        if len(operation_as_list) != len(other_operation_as_list):
            print('\t[X]Operation difference due to length.')
            return False
        
        #Operations are same length, zip is basically parallel comparison
        for row, other_row in zip(operation_as_list, other_operation_as_list):
            if row != other_row:
                print('\t[X]Operation difference due to value.')
                return False
            
        return True
    
    '''
    returns true if the current operation has no rows
    @return: True if the operation is empty; False otherwise
    '''
    def is_empty(self):
        return len(self.rows) == 0
    
    '''
    adds a row to the operation,
        detects for None value types (rather than ''),
        ensures the row is now all empty cells
    @param row: row from openpyxl ReadOnlyWorksheet
    '''
    def add_row(self, row):
        if not all(cell.value == None or cell.value == '' for cell in row):
            if any(cell.value == None for cell in row):
                pass
                #print('Detected None cell value.')
            try:
                if any(cell.fill.fgColor.rgb == RED for cell in row):
                    self.flagged_for_removal = True
                elif any(part_of_change(cell) == True for cell in row):
                    self.part_of_change = True
            except AttributeError:
                #When cells have fill type 'None' program crashes due to 'None' not having a value 'RGB'
                pass
            self.rows.append([cell for cell in row])
            
    '''
    prints the operation to the console
    @param log: log shall be set to None unless called by log_operation
    '''
    def print_operation(self, log=None):
        for row in self.rows:
            print([cell.value for cell in row], file=log)
    
    '''
    logs the operation to the given file
        append mode - if file exists it will be appended to not overwritten
    @param file_name: file_name of the log file
    '''
    def log_operation(self, file_name):
        with open(file_name, 'a+') as log_file:
            self.print_operation(log=log_file)
        log_file.close()
    
    '''
    sets and cleans the operation number incoming from a openpyxl ReadOnlyCell
        needs to cast to string in case Excel converted operation number to float
        ensures the format XXXX.XXXX 
        note that 100.XXXX and 10.XXXX are legal operation numbers
    @param operation_no: incoming cell value from where the operation number was encountered
    '''
    def fix_operation_no(self, operation_no):
        operation_no = str(operation_no)
        
        if '.' not in operation_no:
            operation_no = operation_no + '.0000'
        else:
            while len(operation_no.split('.')[1]) < 4:
                operation_no  = operation_no + '0'
        return operation_no
    
    '''
    returns the operation number to the caller
    @return: returns the operation's operation number to caller as float
    '''
    def get_operation_no(self):
        return float(self.fix_operation_no(self.rows[0][1].value))
    
    '''
    returns the operation number to the caller as a string
    @return: returns a str representation of the operations operation number
    '''
    def get_operation_no_str(self):
        return self.fix_operation_no(self.rows[0][1].value)
    
    '''
    returns an array of openpyxl WriteOnlyCells that can than be outputted to the diff worksheet
        depending on parameters, the caller will recieve properly highlighted cells
    @param route_type: RTE and SM are valid types. Will append placeholder WriteOnlyCells appropriately
        RTE will append cells in the order (RTE value, blank)
        SM will append cells in the order (blank, SM value)
    @param extra: if operation is not in other route extra is True, false otherwise
    @param difftype: If difference is inside scope of change, difftpye will be orange,
        blue otherwise
    '''
    #Theres probably a really short way to do this with list comprehension...
    def get_operation_as_output(self, route_type, extra=False, difftype=None):
        #2D array to return
        returned = []
        
        
        for row in self.rows:
            #temp is a temporary list that will act as a spreadsheet 'row'
            temp = []
            for cell in row:
                
                if route_type == 'RTE':
                    #If the current RTE cell is an extra operation, highlight it as a difference
                    if(extra):
                        temp.append(create_cell(cell.value, 'RTE', diff=difftype))
                    #The current RTE cell does not need any highlighting
                    else:
                        temp.append(create_cell(cell.value, 'RTE'))
                        
                    #Add blank SM cell (placeholder)
                    temp.append(create_cell('', 'SM'))
                
                elif route_type == 'SM':
                    #Add blank RTE cell (placeholder)
                    temp.append(create_cell('', 'RTE'))
                    
                    #If the current SM cell is an extra operation, highlight it as a difference
                    if(extra):
                        temp.append(create_cell(cell.value, 'SM', diff=difftype))
                    #Current SM cell does not need any highlighting
                    else:
                        temp.append(create_cell(cell.value, 'SM'))
            #append the 'row'
            returned.append(temp)
        return returned
    
    '''
    returns the operation as a cleaned* 2D list to the caller of plain string type values
        *All none types will be convered into empty strings for ease of comparison
        *Comment column will be returned as ''
            this is done by enumerating the cell and row values which will act as a row,column counter
            if row_count and cell_count are 0, we know its the comment area
            [0,0] [0,1] [0,2]
            [1,0] [1,1] [1,2]
            [2,0] [2,1] [2,2]
    @return: returns the operation represented as a 2D list (think of slice of a spreadsheet)
    '''
    
    def get_operation_as_list(self):
        return [['' if cell_count == 0 and row_count == 0 else cell.value if cell.value != None else '' for cell_count, cell in enumerate(row)] for row_count, row in enumerate(self.rows)]
    '''
    list comprehension equivalent to:
    returned = []
    for row in self.rows:
        temp = []
        for cell in row:
            if cell.value != None:
                temp.append(cell.value)
            else:
                temp.append('')
    return returned
    <3 python
    '''

def format_output(filepath):
    
    book = oxl.load_workbook(filepath)
    
    
    sheet = book['RTE-SM Compare'] 
    
    sheet.row_dimensions[1].height = 55
    
    
    header_list = [
        'Comments',
        'Full Oper Num',
        'Module',
        'Module Description',
        'Process Def',
        'Photo Layer',
        'PD Type',
        'PD Name',
        'PD Description',
        'Department',
        'PD User Data Sets - Name',
        'PD User Data Sets - Value',
        'Operation User Data Sets - Name',
        'Operation User Data Sets - Value',
        'Mondatory',
        "Carrier Category",
        "LR Context Type",
        "LR Context",
        "Logical Recipe",
        "LR Description",
        "Resolved Pre1 Script",
        "Resolved Pre2 Script",
        'Resolved Post Script',
        'Equipment',
        'Equipment Recipe',
        'Recipe Description',
        'Stage ID',
        'Proc (min)',
        'Wait (min)',
        'WPH'
    ]
    for col in range(1,61,2):
        #merge necessary cells
        sheet.merge_cells(start_row=1, end_row=1, start_column=col, end_column=col+1)
        sheet.merge_cells(start_row=2, end_row=2, start_column=col, end_column=col+1)
        sheet.merge_cells(start_row=3, end_row=3, start_column=col, end_column=col+1)
        
        #assign values, center align
        sheet.cell(row=1, column=col, value=header_list[int((col-1)/2)]).alignment = oxl.styles.alignment.Alignment(horizontal='center', wrap_text=True)
        sheet.cell(row=2, column=col, value="Orange: ##").alignment = oxl.styles.alignment.Alignment(horizontal='center', wrap_text=True)
        sheet.cell(row=3, column=col, value="Blue: ##").alignment = oxl.styles.alignment.Alignment(horizontal='center', wrap_text=True)
        
        #cell coloring
        sheet.cell(row=1, column=col).fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor='FFC0C0C0')
        sheet.cell(row=2, column=col).fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor='FFFABF8F')
        sheet.cell(row=3, column=col).fill = oxl.styles.fills.PatternFill(patternType='solid', fgColor='FF92CDDC')
        
        #column sizing
        sheet.column_dimensions[oxl.utils.get_column_letter(col)].width = 15
        sheet.column_dimensions[oxl.utils.get_column_letter(col+1)].width = 15
        
        #borders
        sheet.cell(row=1, column=col).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        sheet.cell(row=1, column=col+1).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        sheet.cell(row=2, column=col).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        sheet.cell(row=2, column=col+1).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        sheet.cell(row=3, column=col).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        sheet.cell(row=3, column=col+1).border = Border(left=Side(style='thin'),
                                                      right=Side(style='thin'),
                                                      top=Side(style='thin'),
                                                      bottom=Side(style='thin'))
        
        #font
        sheet.cell(row=1, column=col).font = oxl.styles.fonts.Font(name='Consolas',
                                                                  bold=True,
                                                                  size=14)
        sheet.cell(row=2, column=col).font = oxl.styles.fonts.Font(name='Consolas',
                                                                  bold=True,
                                                                  size=10)
        sheet.cell(row=3, column=col).font = oxl.styles.fonts.Font(name='Consolas',
                                                                  bold=True,
                                                                  size=10)
    blue = 0
    orange = 0
    #loop through each column in the sheet
    for index, col in enumerate(sheet.columns):
        
        #loop through each cell in each column
        for cell in col:
            #detecting and counting orange color
            #print(cell.fill.fgColor.rgb)
            if (cell.fill.fgColor.rgb == ORANGE_DIFF_HEX):
                orange += 1
            #detecting and counting blue color
            if(cell.fill.fgColor.rgb == BLUE_DIFF_HEX):
                blue += 1
            #orange counter located at row 2
            #blue counter located at row 3
        if index % 2 != 0:
            currcol = oxl.utils.coordinate_from_string(prev.coordinate)[0]
            orange_coord = currcol + '2'
            blue_coord = currcol + '3'
            sheet[orange_coord].value = "Orange: " + str(orange) 
            sheet[blue_coord].value = "Blue: " + str(blue)
            #print(orange_coord, blue_coord,sheet[orange_coord].value, sheet[blue_coord].value, sheet[blue_coord])
            #reset counters
            blue = 0
            orange = 0
        prev = cell
        #print(currcol)
    sheet.sheet_view.zoomScale = 70
    book.save(filepath)

'''
writes the header to the openpyxl WriteOnly spreadsheet output_worksheet
'''
def write_header():
    rte_header = WriteOnlyCell(ws=output_worksheet, value='RTE')
    sm_header = WriteOnlyCell(ws=output_worksheet, value='SM')
    
    rte_header.fill = RTE_FILL
    sm_header.fill = SM_FILL
    
    rte_header.border = ALL_BORDER
    sm_header.border = ALL_BORDER
    
    rte_header.font = HEADER_FONT
    sm_header.font = HEADER_FONT
    
    rte_header.alignment = CENTER_ALIGN
    sm_header.alignment = CENTER_ALIGN
    
    temp = []
    for header_cell in range(NUM_OUTPUT_COLS):
        temp.append(rte_header)
        temp.append(sm_header)
    output_worksheet.append(temp)

'''
High-level program logic
'''
def load_route(sheetname):
    '''
    @sheet: 
    '''
    sheet = oxl_workbook[sheetname]

    if sheet.max_row <= 10 or sheet.max_column <= 10:
        return False
    
    '''
    @report_id:
    @start_reading:
    @start_index:
    '''
    report_id = ''
    start_reading = False
    start_index = -1
    
    '''
    @loop:
    @operation:
    @route:
    '''
    operation = Operation()
    route = Route()
    for index, row in enumerate(sheet.rows):
        
        #Get flow report header store in report_id
        if 'Flow Report' in stringify(row[COL_A]):
            report_id = stringify(row[COL_A])
            
        #If we encounter 'Full Oper Num', set the start index to two rows later
        if stringify(row[COL_B]) == OPERATION_START:
            start_index = index + 2
        
        #First row of operations, set flag to begin reading data
        if index == start_index:
            start_reading = True
            
        #'Read data mode'
        if start_reading:
            
            #If the current operation list isn't empty (accounting for first loop iteration)
            #and the current cell contains an operation number
            #store the operation in the route, clear out temp operation
            if has_value(row[COL_B]) and not operation.is_empty():
                if not operation.flagged_for_removal:
                    route.add_operation(operation)
                else:
                    print('Operation flagged for removal. Ignoring operation:', operation)
                operation = Operation()               
                
            #If the current row is not all empty cells, add it to the current operation
            operation.add_row(row)
                
    #Last operation will still be stored but not added
    if not operation.is_empty():
        route.add_operation(operation)

    if sheetname == RTE_SHEET and RTE_VALIDATION not in report_id:
        return False
    elif sheetname == SM_SHEET and SM_VALIDATION not in report_id:
        return False
    elif sheetname == MM_SHEET and MM_VALIDATION not in report_id:
        return False
        
    #Set route attributes
    route.set_route_id(get_route_id(report_id))
    route.set_product_id(get_product_id(report_id))
    route.set_route_type(sheetname)
    
    return route


# 'main' function area

#mainbook = xw.Book.caller()


#********************************
# Lines are to be enabled for Excel testing and deployment!
#********************************
mainbook = xw.Book.caller()
oxl_workbook = oxl.load_workbook(mainbook.fullname, read_only=True)

'''
mainbook = xw.Book(WORKBOOK_PATH)
oxl_workbook = oxl.load_workbook(WORKBOOK_PATH, read_only=True)
'''
homesheet = mainbook.sheets[HOME_SHEET]
homesheet.range('A13').value = mainbook.fullname

if homesheet.range('I1').value:
    mmsheet = True
else:
    mmsheet = False

#load routes
homesheet.range('A3:A12').value = ''
homesheet.range('A3:A12').color = HOME_BLUE_RGB


homesheet.range('A3').value = 'Now loading routes.'

rte_route = load_route(RTE_SHEET)
homesheet.range('A4').value = 'Loaded RTE route.'
homesheet.range('A9').value = str(rte_route)

sm_route = load_route(SM_SHEET)
homesheet.range('A5').value = 'Loaded SM route.'
homesheet.range('A10').value = str(sm_route)

if mmsheet:
    mm_route = load_route(MM_SHEET)
    homesheet.range('A6').value = 'Loaded MM route.'
    homesheet.range('A11').value = str(mm_route)
else:
    mm_route = None
    homesheet.range('A11').value = 'No MM metadata.'
    sheet = oxl_workbook[MM_SHEET]

    if sheet.max_row > 1 or sheet.max_column > 1:
        homesheet.range('A6').value = 'WARNING: MM Data detected, but MM Sheet is not checked.'
        homesheet.range('A6').color = ORANGE_WARNING_RGB
    else:
        homesheet.range('A6').value = 'No MM Sheet read.'


if rte_route == False:
    homesheet.range('A4').value = 'INVALID RTE SHEET'
    homesheet.range('A4').color = RED_RGB
if sm_route == False:
    homesheet.range('A5').value = 'INVALID SM SHEET'
    homesheet.range('A5').color = RED_RGB
if mm_route == False:
    homesheet.range('A6').value = 'INVALID MM SHEET'
    homesheet.range('A6').color = RED_RGB






#compare routes
if any (flag == False for flag in [rte_route, sm_route, mm_route]):
    homesheet.range('A7').value = 'UNABLE TO COMPARE ROUTES.'
    homesheet.range('A7').color = RED_RGB
    homesheet.range('A8').value = 'Please check above cells identify error.'
else:
    #prep output
    output_workbook = init_write_only_output()
    output_worksheet = output_workbook[OUTPUT_TAB]  
    homesheet.range('A7').value = 'Now comparing routes...'
    compare_routes(rte_route, sm_route, mm_route)
    homesheet.range('A8').value = 'Comparison completed. Now saving report...'
    print(get_output_options())
    save_output()
    #save output

print('Complete.')
homesheet.range('A12').value = 'Complete.'